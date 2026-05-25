from __future__ import annotations

from dataclasses import dataclass, replace
from datetime import date, datetime
from pathlib import Path
from typing import Any
import warnings

from openpyxl import load_workbook

from app.config.data_conventions import CONTEXT_VALUES, L1_CODES, SPEAKER_TYPES, STANDARD_VARIETIES, build_session_id_from_ref, normalize_session_ref
from language_config import resolve_language_config


ACTIVE_SHEETS: tuple[str, ...] = (
    "Secure_Person_Intake",
    "Research_Person",
    "Research_Session_Intake",
    "Exposure",
    "Vocabularies",
)

EXPOSURE_TYPES: tuple[str, ...] = (
    "study",
    "erasmus",
    "work",
    "travel",
    "family",
    "volunteering",
    "school_exchange",
    "other",
    "unknown",
)

STANDARD_VARIETY_ALIASES: dict[str, str] = {
    "CH_FR_STD": "fr_ch_std",
    "FR_CH_STD": "fr_ch_std",
    "CH_DE_STD": "de_ch_std",
    "DE_CH_STD": "de_ch_std",
}


@dataclass(frozen=True, slots=True)
class SecurePersonIntakeRow:
    person_id: str
    research_consent_signed: str | None
    teaching_consent_signed: str | None
    consent_date: date | None
    consent_file: str | None
    questionnaire_file: str | None
    secure_notes: str | None
    row_number: int


class IntakeWorkbookError(ValueError):
    """Raised when the intake workbook violates the active contract."""


@dataclass(frozen=True, slots=True)
class SessionLinkKey:
    person_id: str
    session_ref: str


@dataclass(frozen=True, slots=True)
class IntakePersonRow:
    person_id: str
    speaker_type: str
    l1: str | None
    l1_additional: tuple[str, ...]
    mother_l1: str | None
    father_l1: str | None
    additional_languages: tuple[str, ...]
    gender: str | None
    birth_year: int | None
    current_region: str | None
    childhood_region: str | None
    origin_country: str | None
    origin_region: str | None
    needs_review: bool
    person_notes: str | None
    row_number: int
    research_consent_signed: str | None = None
    teaching_consent_signed: str | None = None
    consent_date: date | None = None
    consent_file: str | None = None
    questionnaire_file: str | None = None
    secure_notes: str | None = None


@dataclass(frozen=True, slots=True)
class IntakeSessionRow:
    person_id: str
    session_ref: str
    session_id: str
    target_language: str
    corpus_language: str
    standard_variety: str | None
    level_self: str | None
    level_code: str | None
    recording_year: int
    recording_date: date | None
    recorded_by: str | None
    context: str | None
    needs_review: bool
    session_notes: str | None
    row_number: int


@dataclass(frozen=True, slots=True)
class IntakeExposureRow:
    person_id: str
    session_ref: str
    target_language: str
    country: str | None
    duration_months: float | None
    exposure_type: str | None
    exposure_notes: str | None
    needs_review: bool
    row_number: int


@dataclass(frozen=True, slots=True)
class IntakeWorkbookData:
    workbook_path: Path
    target_language: str
    persons: dict[str, IntakePersonRow]
    sessions: tuple[IntakeSessionRow, ...]
    exposures_by_key: dict[SessionLinkKey, tuple[IntakeExposureRow, ...]]
    warnings: tuple[str, ...]
    errors: tuple[str, ...]


def _normalize_optional_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        normalized = value.strip()
        return normalized or None
    normalized = str(value).strip()
    return normalized or None


def _normalize_bool_flag(value: Any, *, field_name: str, row_number: int) -> bool:
    text = (_normalize_optional_text(value) or "no").lower()
    if text in {"no", "nein", "false", "0"}:
        return False
    if text in {"yes", "ja", "true", "1"}:
        return True
    raise IntakeWorkbookError(f"Invalid {field_name} value at row {row_number}: expected yes/no, got {value!r}")


def _normalize_int(value: Any, *, field_name: str, row_number: int, required: bool = False) -> int | None:
    if value in (None, ""):
        if required:
            raise IntakeWorkbookError(f"Missing required {field_name} at row {row_number}")
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    text = _normalize_optional_text(value)
    if text and text.isdigit():
        return int(text)
    raise IntakeWorkbookError(f"Invalid {field_name} value at row {row_number}: expected integer, got {value!r}")


def _normalize_yes_no_unknown(value: Any, *, field_name: str, row_number: int) -> str | None:
    text = (_normalize_optional_text(value) or "").lower()
    if not text:
        return None
    if text in {"yes", "no", "unknown"}:
        return text
    raise IntakeWorkbookError(
        f"Invalid {field_name} value at row {row_number}: expected yes/no/unknown or empty, got {value!r}"
    )


def _normalize_duration_months(value: Any, *, row_number: int) -> tuple[float | None, str | None]:
    if value in (None, ""):
        return None, None
    if isinstance(value, bool):
        return None, f"Invalid duration_months value in Exposure row {row_number}: expected numeric months, got {value!r}; field left empty."
    if isinstance(value, (int, float)):
        return float(value), None
    text = _normalize_optional_text(value)
    if text is None:
        return None, None
    normalized = text.replace(",", ".")
    try:
        return float(normalized), None
    except ValueError:
        return None, f"Invalid duration_months value in Exposure row {row_number}: expected numeric months, got {value!r}; field left empty."


def _normalize_date(value: Any, *, field_name: str, row_number: int) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _normalize_optional_text(value)
    if text is None:
        return None
    try:
        return date.fromisoformat(text.split(" ", 1)[0])
    except ValueError as exc:
        raise IntakeWorkbookError(f"Invalid {field_name} value at row {row_number}: expected ISO date, got {value!r}") from exc


def _normalize_multivalue_text(value: Any) -> tuple[str, ...]:
    text = _normalize_optional_text(value)
    if text is None:
        return tuple()
    return tuple(part.strip() for part in text.split(";") if part and part.strip())


def _normalize_l1_field(value: Any, *, row_number: int, field_name: str, allow_unknown: bool = False) -> str | None:
    text = _normalize_optional_text(value)
    if text is None:
        return None
    if allow_unknown and text.lower() == "unknown":
        return "unknown"
    normalized = text.upper()
    if normalized not in L1_CODES:
        raise IntakeWorkbookError(
            f"Invalid {field_name} value at row {row_number}: expected one of {', '.join(L1_CODES)} or empty, got {value!r}"
        )
    return normalized


def _normalize_l1_additional(value: Any, *, row_number: int) -> tuple[str, ...]:
    values = _normalize_multivalue_text(value)
    normalized_values: list[str] = []
    seen: set[str] = set()
    for entry in values:
        normalized = entry.upper()
        if normalized not in L1_CODES:
            raise IntakeWorkbookError(
                f"Invalid l1_additional value at row {row_number}: expected one of {', '.join(L1_CODES)}, got {entry!r}"
            )
        if normalized in seen:
            continue
        normalized_values.append(normalized)
        seen.add(normalized)
    return tuple(normalized_values)


def _normalize_target_language(value: Any, *, field_name: str, row_number: int) -> str:
    text = _normalize_optional_text(value)
    if text is None:
        raise IntakeWorkbookError(f"Missing {field_name} at row {row_number}")
    return resolve_language_config(text).code


def _row_dict(headers: tuple[str, ...], values: tuple[Any, ...]) -> dict[str, Any]:
    return {header: values[index] if index < len(values) else None for index, header in enumerate(headers)}


def _normalize_standard_variety(value: Any, *, row_number: int, target_language: str) -> str | None:
    text = _normalize_optional_text(value)
    if text is None:
        return None
    workbook_value = text.upper()
    normalized = STANDARD_VARIETY_ALIASES.get(workbook_value, workbook_value.lower())
    if normalized not in STANDARD_VARIETIES[target_language]:
        raise IntakeWorkbookError(
            f"Invalid standard_variety in Research_Session_Intake row {row_number}: {value!r}"
        )
    return normalized


def _normalize_exposure_type(value: Any, *, row_number: int) -> tuple[str | None, str | None]:
    normalized = (_normalize_optional_text(value) or "").lower()
    if not normalized:
        return None, None
    if normalized == "unspecified":
        return "unknown", f"Deprecated exposure type unspecified in Exposure row {row_number}; normalized to unknown."
    if normalized == "study_abroad":
        return "study", f"Deprecated exposure type study_abroad in Exposure row {row_number}; normalized to study."
    if normalized not in EXPOSURE_TYPES:
        raise IntakeWorkbookError(
            f"Invalid type value in Exposure row {row_number}: expected one of {', '.join(EXPOSURE_TYPES)}, got {value!r}"
        )
    return normalized, None


def _validate_required_headers(
    sheet_name: str,
    headers: tuple[str, ...],
    *,
    warnings_out: list[str],
) -> tuple[str, ...]:
    normalized_headers = headers
    if sheet_name == "Secure_Person_Intake" and "research_consent_signed" not in normalized_headers and "consent_signed" in normalized_headers:
        normalized_headers = tuple("research_consent_signed" if header == "consent_signed" else header for header in normalized_headers)
        warnings_out.append("Deprecated column consent_signed used; please rename to research_consent_signed.")
    if sheet_name == "Vocabularies":
        normalized_headers = tuple("l1_code" if header.startswith("l1_code") else header for header in normalized_headers)

    required_headers: dict[str, tuple[str, ...]] = {
        "Secure_Person_Intake": (
            "person_id",
            "last_name",
            "first_name",
            "email",
            "research_consent_signed",
            "consent_date",
            "consent_file",
            "teaching_consent_signed",
            "questionnaire_file",
            "paper_original_location",
            "intake_date",
            "intake_by",
            "needs_review",
            "verified_by",
            "verified_date",
            "secure_notes",
        ),
        "Research_Person": (
            "person_id",
            "speaker_type",
            "l1",
            "l1_additional",
            "mother_l1",
            "father_l1",
            "additional_languages",
            "gender",
            "birth_year",
            "current_region",
            "childhood_region",
            "origin_country",
            "origin_region",
            "needs_review",
            "person_notes",
        ),
        "Research_Session_Intake": (
            "person_id",
            "session_ref",
            "session_id",
            "target_language",
            "standard_variety",
            "level_self",
            "level_code",
            "recording_year",
            "recording_date",
            "recorded_by",
            "context",
            "needs_review",
            "session_notes",
        ),
        "Exposure": (
            "person_id",
            "session_ref",
            "target_language",
            "country",
            "duration_months",
            "type",
            "exposure_notes",
            "needs_review",
        ),
        "Vocabularies": (
            "gender",
            "speaker_type",
            "l1_code",
            "target_language",
            "level_code",
            "level_self",
            "standard_variety",
            "context",
            "exposure_type",
            "task_type",
            "recorded_by",
            "yes_no_unknown",
        ),
    }
    missing = [header for header in required_headers[sheet_name] if header not in normalized_headers]
    if missing:
        raise IntakeWorkbookError(f"Workbook sheet {sheet_name!r} is missing required headers: {', '.join(missing)}")
    return normalized_headers


def _normalize_person_row(row: dict[str, Any], row_number: int) -> IntakePersonRow:
    person_id = (_normalize_optional_text(row.get("person_id")) or "").upper()
    if not person_id:
        raise IntakeWorkbookError(f"Missing person_id in Research_Person row {row_number}")

    speaker_type = (_normalize_optional_text(row.get("speaker_type")) or "").lower()
    if speaker_type not in SPEAKER_TYPES:
        raise IntakeWorkbookError(f"Invalid speaker_type in Research_Person row {row_number}: {row.get('speaker_type')!r}")

    return IntakePersonRow(
        person_id=person_id,
        speaker_type=speaker_type,
        l1=_normalize_l1_field(row.get("l1"), row_number=row_number, field_name="l1"),
        l1_additional=_normalize_l1_additional(row.get("l1_additional"), row_number=row_number),
        mother_l1=_normalize_l1_field(row.get("mother_l1"), row_number=row_number, field_name="mother_l1", allow_unknown=True),
        father_l1=_normalize_l1_field(row.get("father_l1"), row_number=row_number, field_name="father_l1", allow_unknown=True),
        additional_languages=_normalize_multivalue_text(row.get("additional_languages")),
        gender=(_normalize_optional_text(row.get("gender")) or "").lower() or None,
        birth_year=_normalize_int(row.get("birth_year"), field_name="birth_year", row_number=row_number),
        current_region=_normalize_optional_text(row.get("current_region")),
        childhood_region=_normalize_optional_text(row.get("childhood_region")),
        origin_country=_normalize_optional_text(row.get("origin_country")),
        origin_region=_normalize_optional_text(row.get("origin_region")),
        needs_review=_normalize_bool_flag(row.get("needs_review"), field_name="needs_review", row_number=row_number),
        person_notes=_normalize_optional_text(row.get("person_notes")),
        row_number=row_number,
    )


def _normalize_secure_person_row(row: dict[str, Any], row_number: int) -> SecurePersonIntakeRow:
    person_id = (_normalize_optional_text(row.get("person_id")) or "").upper()
    if not person_id:
        raise IntakeWorkbookError(f"Missing person_id in Secure_Person_Intake row {row_number}")
    return SecurePersonIntakeRow(
        person_id=person_id,
        research_consent_signed=_normalize_yes_no_unknown(
            row.get("research_consent_signed"),
            field_name="research_consent_signed",
            row_number=row_number,
        ),
        teaching_consent_signed=_normalize_yes_no_unknown(
            row.get("teaching_consent_signed"),
            field_name="teaching_consent_signed",
            row_number=row_number,
        ),
        consent_date=_normalize_date(row.get("consent_date"), field_name="consent_date", row_number=row_number),
        consent_file=_normalize_optional_text(row.get("consent_file")),
        questionnaire_file=_normalize_optional_text(row.get("questionnaire_file")),
        secure_notes=_normalize_optional_text(row.get("secure_notes")),
        row_number=row_number,
    )


def _raw_person_id(value: Any) -> str | None:
    person_id = (_normalize_optional_text(value) or "").upper()
    return person_id or None


def _normalize_session_row(row: dict[str, Any], row_number: int) -> tuple[str | None, IntakeSessionRow | None, str | None]:
    person_id = (_normalize_optional_text(row.get("person_id")) or "").upper()
    if not person_id:
        raise IntakeWorkbookError(f"Missing person_id in Research_Session_Intake row {row_number}")

    session_ref = normalize_session_ref(row.get("session_ref"))
    if session_ref is None:
        raise IntakeWorkbookError(f"Invalid session_ref in Research_Session_Intake row {row_number}: {row.get('session_ref')!r}")

    target_language = _normalize_target_language(
        row.get("target_language"),
        field_name="target_language",
        row_number=row_number,
    )
    language_config = resolve_language_config(target_language)

    existing_session_id = _normalize_optional_text(row.get("session_id"))
    warning = None
    if existing_session_id:
        warning = f"Research_Session_Intake row {row_number} contains session_id={existing_session_id!r}; ignored because session_id is derived."

    recording_year = _normalize_int(row.get("recording_year"), field_name="recording_year", row_number=row_number, required=True)
    if recording_year is None:
        raise IntakeWorkbookError(f"Missing recording_year in Research_Session_Intake row {row_number}")

    level_self = _normalize_optional_text(row.get("level_self"))
    level_code = _normalize_optional_text(row.get("level_code"))
    standard_variety = _normalize_standard_variety(row.get("standard_variety"), row_number=row_number, target_language=target_language)
    context = (_normalize_optional_text(row.get("context")) or "").lower() or None
    if context is not None and context not in CONTEXT_VALUES:
        raise IntakeWorkbookError(f"Invalid context in Research_Session_Intake row {row_number}: {row.get('context')!r}")

    try:
        session_id = build_session_id_from_ref(person_id, recording_year, session_ref)
    except ValueError as exc:
        raise IntakeWorkbookError(f"Invalid person/session combination in Research_Session_Intake row {row_number}: {exc}") from exc

    session = IntakeSessionRow(
        person_id=person_id,
        session_ref=session_ref,
        session_id=session_id,
        target_language=target_language,
        corpus_language=language_config.corpus_slug,
        standard_variety=standard_variety,
        level_self=level_self,
        level_code=level_code,
        recording_year=recording_year,
        recording_date=_normalize_date(row.get("recording_date"), field_name="recording_date", row_number=row_number),
        recorded_by=_normalize_optional_text(row.get("recorded_by")),
        context=context,
        needs_review=_normalize_bool_flag(row.get("needs_review"), field_name="needs_review", row_number=row_number),
        session_notes=_normalize_optional_text(row.get("session_notes")),
        row_number=row_number,
    )
    return target_language, session, warning


def _normalize_exposure_row(row: dict[str, Any], row_number: int) -> tuple[str, IntakeExposureRow, tuple[str, ...]]:
    person_id = (_normalize_optional_text(row.get("person_id")) or "").upper()
    if not person_id:
        raise IntakeWorkbookError(f"Missing person_id in Exposure row {row_number}")
    session_ref = normalize_session_ref(row.get("session_ref"))
    if session_ref is None:
        raise IntakeWorkbookError(f"Invalid session_ref in Exposure row {row_number}: {row.get('session_ref')!r}")
    target_language = _normalize_target_language(row.get("target_language"), field_name="target_language", row_number=row_number)
    duration_months, duration_warning = _normalize_duration_months(row.get("duration_months"), row_number=row_number)
    exposure_type, type_warning = _normalize_exposure_type(row.get("type"), row_number=row_number)
    warnings_out = tuple(warning for warning in (duration_warning, type_warning) if warning is not None)
    exposure = IntakeExposureRow(
        person_id=person_id,
        session_ref=session_ref,
        target_language=target_language,
        country=_normalize_optional_text(row.get("country")),
        duration_months=duration_months,
        exposure_type=exposure_type,
        exposure_notes=_normalize_optional_text(row.get("exposure_notes")),
        needs_review=_normalize_bool_flag(row.get("needs_review"), field_name="needs_review", row_number=row_number),
        row_number=row_number,
    )
    return target_language, exposure, warnings_out


def _load_sheet_headers(workbook, sheet_name: str) -> tuple[str, ...]:
    if sheet_name not in workbook.sheetnames:
        raise IntakeWorkbookError(f"Workbook is missing required sheet {sheet_name!r}")
    worksheet = workbook[sheet_name]
    first_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if first_row is None:
        raise IntakeWorkbookError(f"Workbook sheet {sheet_name!r} is empty")
    headers = tuple(str(cell).strip() if cell is not None else "" for cell in first_row)
    if any(not header for header in headers):
        raise IntakeWorkbookError(f"Workbook sheet {sheet_name!r} contains empty header cells")
    return headers


def load_intake_workbook(
    workbook_path: Path,
    *,
    target_language: str,
    person_id_filter: str | None = None,
) -> IntakeWorkbookData:
    normalized_target_language = resolve_language_config(target_language).code
    normalized_person_filter = (_normalize_optional_text(person_id_filter) or "").upper() or None

    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message="Data Validation extension is not supported and will be removed",
            category=UserWarning,
        )
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        try:
            warnings_out: list[str] = []
            errors_out: list[str] = []

            for sheet_name in ACTIVE_SHEETS:
                if sheet_name not in workbook.sheetnames:
                    raise IntakeWorkbookError(f"Workbook is missing required sheet {sheet_name!r}")

            secure_headers = _validate_required_headers(
                "Secure_Person_Intake",
                _load_sheet_headers(workbook, "Secure_Person_Intake"),
                warnings_out=warnings_out,
            )
            person_headers = _validate_required_headers(
                "Research_Person",
                _load_sheet_headers(workbook, "Research_Person"),
                warnings_out=warnings_out,
            )
            session_headers = _validate_required_headers(
                "Research_Session_Intake",
                _load_sheet_headers(workbook, "Research_Session_Intake"),
                warnings_out=warnings_out,
            )
            exposure_headers = _validate_required_headers(
                "Exposure",
                _load_sheet_headers(workbook, "Exposure"),
                warnings_out=warnings_out,
            )
            _validate_required_headers(
                "Vocabularies",
                _load_sheet_headers(workbook, "Vocabularies"),
                warnings_out=warnings_out,
            )

            filtered_sessions: list[IntakeSessionRow] = []
            session_keys: set[SessionLinkKey] = set()
            referenced_person_ids: set[str] = set()
            session_sheet = workbook["Research_Session_Intake"]
            for row_number, row_values in enumerate(session_sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not any(value not in (None, "") for value in row_values):
                    continue
                row = _row_dict(session_headers, row_values)
                try:
                    row_target_language = _normalize_target_language(
                        row.get("target_language"),
                        field_name="target_language",
                        row_number=row_number,
                    )
                except IntakeWorkbookError as exc:
                    errors_out.append(str(exc))
                    continue
                if row_target_language != normalized_target_language:
                    continue
                try:
                    session_target_language, session, session_warning = _normalize_session_row(row, row_number)
                except IntakeWorkbookError as exc:
                    errors_out.append(str(exc))
                    continue
                if session_warning is not None:
                    warnings_out.append(session_warning)
                if session_target_language != normalized_target_language:
                    continue
                if normalized_person_filter is not None and session.person_id != normalized_person_filter:
                    continue
                key = SessionLinkKey(person_id=session.person_id, session_ref=session.session_ref)
                if key in session_keys:
                    errors_out.append(f"Duplicate Research_Session_Intake link for {session.person_id}/{session.session_ref} (row {row_number})")
                    continue
                session_keys.add(key)
                referenced_person_ids.add(session.person_id)
                filtered_sessions.append(session)

            all_people: dict[str, IntakePersonRow] = {}
            person_sheet = workbook["Research_Person"]
            for row_number, row_values in enumerate(person_sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not any(value not in (None, "") for value in row_values):
                    continue
                row = _row_dict(person_headers, row_values)
                raw_person_id = _raw_person_id(row.get("person_id"))
                if raw_person_id is None:
                    errors_out.append(f"Missing person_id in Research_Person row {row_number}")
                    continue
                if raw_person_id not in referenced_person_ids:
                    continue
                try:
                    person = _normalize_person_row(row, row_number)
                except IntakeWorkbookError as exc:
                    errors_out.append(str(exc))
                    continue
                if person.person_id in all_people:
                    errors_out.append(f"Duplicate person_id in Research_Person: {person.person_id} (row {row_number})")
                    continue
                all_people[person.person_id] = person

            secure_people: dict[str, SecurePersonIntakeRow] = {}
            secure_sheet = workbook["Secure_Person_Intake"]
            for row_number, row_values in enumerate(secure_sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not any(value not in (None, "") for value in row_values):
                    continue
                row = _row_dict(secure_headers, row_values)
                raw_person_id = _raw_person_id(row.get("person_id"))
                if raw_person_id is None:
                    errors_out.append(f"Missing person_id in Secure_Person_Intake row {row_number}")
                    continue
                if raw_person_id not in referenced_person_ids:
                    continue
                try:
                    secure_person = _normalize_secure_person_row(row, row_number)
                except IntakeWorkbookError as exc:
                    errors_out.append(str(exc))
                    continue
                if secure_person.person_id in secure_people:
                    errors_out.append(
                        f"Duplicate person_id in Secure_Person_Intake: {secure_person.person_id} (row {row_number})"
                    )
                    continue
                secure_people[secure_person.person_id] = secure_person

            filtered_people: dict[str, IntakePersonRow] = {}
            for session in filtered_sessions:
                person = all_people.get(session.person_id)
                if person is None:
                    errors_out.append(
                        f"Research_Session_Intake row {session.row_number} references missing Research_Person {session.person_id}"
                    )
                    continue
                if person.speaker_type == "learner" and session.standard_variety is not None:
                    errors_out.append(
                        f"Learner session row {session.row_number} must not define standard_variety"
                    )
                if person.speaker_type == "native_speaker" and (session.level_self is not None or session.level_code is not None):
                    errors_out.append(
                        f"Native-speaker session row {session.row_number} must not define level_self or level_code"
                    )
                secure_person = secure_people.get(person.person_id)
                if secure_person is None:
                    filtered_people[person.person_id] = person
                    continue
                filtered_people[person.person_id] = replace(
                    person,
                    research_consent_signed=secure_person.research_consent_signed,
                    teaching_consent_signed=secure_person.teaching_consent_signed,
                    consent_date=secure_person.consent_date,
                    consent_file=secure_person.consent_file,
                    questionnaire_file=secure_person.questionnaire_file,
                    secure_notes=secure_person.secure_notes,
                )

            exposures_by_key: dict[SessionLinkKey, list[IntakeExposureRow]] = {}
            exposure_sheet = workbook["Exposure"]
            for row_number, row_values in enumerate(exposure_sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not any(value not in (None, "") for value in row_values):
                    continue
                row = _row_dict(exposure_headers, row_values)
                try:
                    row_target_language = _normalize_target_language(
                        row.get("target_language"),
                        field_name="target_language",
                        row_number=row_number,
                    )
                except IntakeWorkbookError as exc:
                    errors_out.append(str(exc))
                    continue
                if row_target_language != normalized_target_language:
                    continue
                try:
                    exposure_target_language, exposure, exposure_warnings = _normalize_exposure_row(row, row_number)
                except IntakeWorkbookError as exc:
                    errors_out.append(str(exc))
                    continue
                warnings_out.extend(exposure_warnings)
                if exposure_target_language != normalized_target_language:
                    continue
                if normalized_person_filter is not None and exposure.person_id != normalized_person_filter:
                    continue
                key = SessionLinkKey(person_id=exposure.person_id, session_ref=exposure.session_ref)
                if key not in session_keys:
                    warnings_out.append(
                        f"Exposure row {row_number} for {exposure.person_id}/{exposure.session_ref} has no matching imported session and is ignored."
                    )
                    continue
                exposures_by_key.setdefault(key, []).append(exposure)
        finally:
            workbook.close()

    return IntakeWorkbookData(
        workbook_path=workbook_path,
        target_language=normalized_target_language,
        persons=filtered_people,
        sessions=tuple(filtered_sessions),
        exposures_by_key={key: tuple(values) for key, values in exposures_by_key.items()},
        warnings=tuple(warnings_out),
        errors=tuple(errors_out),
    )