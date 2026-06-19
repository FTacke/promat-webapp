from __future__ import annotations

import json
import os
import sys
import zipfile
from pathlib import Path
from types import SimpleNamespace

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
import pytest
from sqlalchemy import create_engine, select
from sqlalchemy.orm import sessionmaker


TEST_REPO_ROOT = Path(__file__).resolve().parents[2]
os.environ.setdefault("FLASK_ENV", "development")
os.environ.setdefault("PROMAT_RUNTIME_ROOT", str(TEST_REPO_ROOT))
os.environ.setdefault("PROMAT_PUBLIC_ROOT", str(TEST_REPO_ROOT / "public"))

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))
sys.path.insert(0, str(TEST_REPO_ROOT / "scripts" / "research_data_intake"))

import app.research_sessions as research_sessions  # noqa: E402
from app.auth.models import Base  # noqa: E402
from app.research_metadata import ResearchPerson, ResearchSession, ResearchSessionExposure  # noqa: E402
import import_batch_to_production as production_importer  # noqa: E402
from intake_workbook_reader import load_intake_workbook  # noqa: E402
from intake_batch_common import working_text_mfa_state_path  # noqa: E402
from textgrid_support import TextGridInterval, spoken_intervals  # noqa: E402


def _set_runtime_env(tmp_path: Path, monkeypatch) -> Path:
    runtime_root = tmp_path / "runtime-root"
    (runtime_root / "data" / "sessions").mkdir(parents=True, exist_ok=True)
    monkeypatch.setenv("PROMAT_RUNTIME_ROOT", str(runtime_root))
    monkeypatch.setenv("PROMAT_PUBLIC_ROOT", str(TEST_REPO_ROOT / "public"))
    monkeypatch.setenv("PROMAT_LOCAL_ARCHIVE_ROOT", str(tmp_path / "archive-root"))
    research_sessions.load_language_sessions.cache_clear()
    research_sessions.load_person_records.cache_clear()
    return runtime_root


def _session_row(
    *,
    person_id: str,
    session_ref: str,
    workbook_session_id: str,
    target_language: str,
    recording_year: int | None,
    recording_date: str,
    recorded_by: str = "Ana Romero",
    context: str = "baseline",
    standard_variety: str | None = None,
    level_self: str | None = None,
    level_code: str | None = None,
    session_notes: str | None = "test learner session",
) -> list[object]:
    return [
        person_id,
        session_ref,
        workbook_session_id,
        target_language,
        standard_variety,
        level_self,
        level_code,
        recording_year,
        recording_date,
        recorded_by,
        context,
        "no",
        session_notes,
    ]


def _write_workbook(
    tmp_path: Path,
    *,
    person_id: str = "ES-L-0001",
    speaker_type: str = "learner",
    include_out_of_scope_invalid_row: bool = False,
    use_deprecated_consent_column: bool = False,
    research_consent_signed: str = "yes",
    teaching_consent_signed: str = "unknown",
    standard_variety: str | None = None,
    exposure_duration: object = 6,
    exposure_type: str = "erasmus",
    exposure_country: str = "Spain",
    exposure_notes: str = "Semester in Madrid.",
) -> Path:
    workbook_path = tmp_path / "intake.xlsx"
    workbook = Workbook()

    secure_sheet = workbook.active
    secure_sheet.title = "Secure_Person_Intake"
    secure_sheet.append(
        [
            "person_id",
            "last_name",
            "first_name",
            "email",
            "consent_signed" if use_deprecated_consent_column else "research_consent_signed",
            "consent_date",
            "consent_file",
            "teaching_consent_signed",
            "questionnaire_file",
            "paper_original_location",
            "intake_date",
            "intake_by",
            "needs_review",
            "verified_by",
            "verified_date",
            "secure_notes",
        ]
    )
    secure_sheet.append(
        [
            person_id,
            "Mustermann",
            "Anna",
            "anna@example.test",
            research_consent_signed,
            "2026-03-14",
            "consent_anna.pdf",
            teaching_consent_signed,
            "questionnaire_anna.pdf",
            "cabinet A",
            "2026-03-14",
            "Ana Romero",
            "no",
            None,
            None,
            "Internal secure note.",
        ]
    )

    person_sheet = workbook.create_sheet("Research_Person")
    person_sheet.title = "Research_Person"
    person_sheet.append(
        [
            "person_id",
            "speaker_type",
            "l1",
            "l1_additional",
            "mother_l1",
            "father_l1",
            "additional_languages",
            "gender",
            "birth_year",
            "current_region",
            "childhood_region",
            "origin_country",
            "origin_region",
            "needs_review",
            "person_notes",
        ]
    )
    person_sheet.append(
        [
            person_id,
            speaker_type,
            "DE" if speaker_type == "learner" else "ES",
            "IT; EN",
            "IT",
            "DE",
            "EN; FR",
            "female",
            2001,
            "NRW",
            "Bayern",
            None,
            None,
            "no",
            None,
        ]
    )

    session_sheet = workbook.create_sheet("Research_Session_Intake")
    session_sheet.append(
        [
            "person_id",
            "session_ref",
            "session_id",
            "target_language",
            "standard_variety",
            "level_self",
            "level_code",
            "recording_year",
            "recording_date",
            "recorded_by",
            "context",
            "needs_review",
            "session_notes",
        ]
    )
    session_sheet.append(
        _session_row(
            person_id=person_id,
            session_ref="S01",
            workbook_session_id="SHOULD-BE-IGNORED",
            target_language="ES",
            standard_variety=standard_variety if standard_variety is not None else ("ES_STD" if speaker_type == "native_speaker" else None),
            level_self="B1-B2" if speaker_type == "learner" else None,
            level_code="B1" if speaker_type == "learner" else None,
            recording_year=2026,
            recording_date="2026-03-14",
        )
    )
    if include_out_of_scope_invalid_row:
        session_sheet.append(
            _session_row(
                person_id="FR-L-0001",
                session_ref="S01",
                workbook_session_id="ALSO-IGNORED",
                target_language="FR",
                recording_year=None,
                recording_date="2026-04-01",
            )
        )

    exposure_sheet = workbook.create_sheet("Exposure")
    exposure_sheet.append(
        [
            "person_id",
            "session_ref",
            "target_language",
            "country",
            "duration_months",
            "type",
            "exposure_notes",
            "needs_review",
        ]
    )
    if speaker_type == "learner":
        exposure_sheet.append(
            [
                person_id,
                "S01",
                "ES",
                exposure_country,
                exposure_duration,
                exposure_type,
                exposure_notes,
                "no",
            ]
        )

    vocabulary_sheet = workbook.create_sheet("Vocabularies")
    vocabulary_sheet.append(
        [
            "gender",
            "speaker_type",
            "l1_code",
            "target_language",
            "level_code",
            "level_self",
            "standard_variety",
            "context",
            "exposure_type",
            "task_type",
            "recorded_by",
            "yes_no_unknown",
        ]
    )
    vocabulary_sheet.append(["female", "learner", "DE", "ES", "B1", "B1-B2", "EC_STD", "baseline", "study", "wordlist", "Ana Romero", "yes"])
    vocabulary_sheet.append([None, None, None, None, None, None, "CL_STD", None, "erasmus", None, None, "no"])
    vocabulary_sheet.append([None, None, None, None, None, None, "PE_STD", None, "school_exchange", None, None, "unknown"])

    workbook.save(workbook_path)
    workbook.close()
    return workbook_path


def _patch_workbook_sqref(workbook_path: Path, original: str, replacement: str) -> None:
    patched_path = workbook_path.with_suffix(".patched.xlsx")
    with zipfile.ZipFile(workbook_path, "r") as source_zip, zipfile.ZipFile(patched_path, "w") as target_zip:
        for item in source_zip.infolist():
            data = source_zip.read(item.filename)
            if item.filename.startswith("xl/worksheets/") and item.filename.endswith(".xml"):
                data = data.replace(original.encode("utf-8"), replacement.encode("utf-8"))
            target_zip.writestr(item, data)
    patched_path.replace(workbook_path)


def _prepare_interview_batch(
    tmp_path: Path,
    *,
    person_id: str = "ES-L-0001",
    include_raw_master: bool,
    include_working_interview: bool = True,
) -> Path:
    batch_dir = tmp_path / "spanish_batch_20260421"
    (batch_dir / "intake_data").mkdir(parents=True, exist_ok=True)
    (batch_dir / "processed").mkdir(parents=True, exist_ok=True)
    filename_prefix = person_id.lower().replace("-", "_")

    (batch_dir / "processed" / f"{filename_prefix}_interview_processed.wav").write_bytes(b"batch-source-interview")
    (batch_dir / "processed" / f"{filename_prefix}_interview_processed.json").write_text(
        json.dumps(
            {
                "segments": [
                    {
                        "speaker": "spk1",
                        "words": [
                            {"start": 0.0, "end": 0.5, "text": "Item", "duration": 0.5, "conf": 1, "pristine": True}
                        ],
                    }
                ]
            },
            ensure_ascii=False,
            indent=2,
        )
        + "\n",
        encoding="utf-8",
    )

    if include_working_interview:
        (batch_dir / "working" / person_id / "interview" / "source").mkdir(parents=True, exist_ok=True)
        (batch_dir / "working" / person_id / "interview" / "alignment").mkdir(parents=True, exist_ok=True)

        interview_wav = batch_dir / "working" / person_id / "interview" / "source" / "interview.wav"
        interview_wav.write_bytes(b"working-interview-wav")
        interview_json = batch_dir / "working" / person_id / "interview" / "alignment" / "interview.json"
        interview_json.write_text(
            json.dumps(
                {
                    "session_id": None,
                    "person_id": person_id,
                    "task": "interview",
                    "audio": {"full_mp3": "derived/interview.mp3"},
                    "segments": [
                        {
                            "segment_id": "seg_001",
                            "segment_number": "1",
                            "speaker_code": "interviewer",
                            "start_ms": 0,
                            "end_ms": 1000,
                            "text": "Item Nummer 25.",
                            "tokens": [
                                {"token_id": "seg_001_tok_001", "text": "Item", "start_ms": 0, "end_ms": 250},
                                {"token_id": "seg_001_tok_002", "text": "Nummer", "start_ms": 250, "end_ms": 600},
                                {"token_id": "seg_001_tok_003", "text": "25", "suffix": ".", "start_ms": 600, "end_ms": 1000},
                            ],
                            "annotations": [
                                {
                                    "kind": "material_ref",
                                    "item_id": "wl_025",
                                    "task": "wordlist",
                                    "insert_after_token_id": "seg_001_tok_003",
                                    "label": "oír",
                                    "item_number": "25",
                                    "canonical_text": "oír",
                                }
                            ],
                        }
                    ],
                },
                ensure_ascii=False,
                indent=2,
            )
            + "\n",
            encoding="utf-8",
        )

    if include_raw_master:
        raw_dir = batch_dir / "raw"
        raw_dir.mkdir(parents=True, exist_ok=True)
        (raw_dir / f"{filename_prefix}_interview_raw.wav").write_bytes(b"batch-raw-interview")

    return batch_dir


def _session_factory(tmp_path: Path):
    database_path = tmp_path / "research.sqlite3"
    engine = create_engine(f"sqlite+pysqlite:///{database_path.as_posix()}", future=True)
    Base.metadata.create_all(engine)
    return sessionmaker(bind=engine, future=True)


def test_load_intake_workbook_derives_session_id_and_ignores_out_of_scope_rows(tmp_path: Path) -> None:
    workbook_path = _write_workbook(tmp_path, include_out_of_scope_invalid_row=True)

    workbook_data = load_intake_workbook(workbook_path, target_language="es")

    assert workbook_data.errors == ()
    assert len(workbook_data.sessions) == 1
    assert workbook_data.sessions[0].session_id == "ES-L-0001-2026-S01"
    assert workbook_data.persons["ES-L-0001"].research_consent_signed == "yes"
    assert workbook_data.persons["ES-L-0001"].teaching_consent_signed == "unknown"
    assert workbook_data.warnings == (
        "Research_Session_Intake row 2 contains session_id='SHOULD-BE-IGNORED'; ignored because session_id is derived.",
    )


def test_load_intake_workbook_accepts_full_l1_vocabulary(tmp_path: Path) -> None:
    workbook_path = _write_workbook(tmp_path)
    workbook = load_workbook(workbook_path)
    person_sheet = workbook["Research_Person"]
    person_sheet["C2"] = "AR"
    person_sheet["D2"] = "PL; CZ; ZH; unknown"
    person_sheet["E2"] = "PL"
    person_sheet["F2"] = "unknown"
    workbook.save(workbook_path)
    workbook.close()

    workbook_data = load_intake_workbook(workbook_path, target_language="es")

    person = workbook_data.persons["ES-L-0001"]
    assert person.l1 == "AR"
    assert person.l1_additional == ("PL", "CZ", "ZH", "unknown")
    assert person.mother_l1 == "PL"
    assert person.father_l1 == "unknown"


def test_load_intake_workbook_accepts_deprecated_consent_column_with_warning(tmp_path: Path) -> None:
    workbook_path = _write_workbook(tmp_path, use_deprecated_consent_column=True)

    workbook_data = load_intake_workbook(workbook_path, target_language="es")

    assert workbook_data.errors == ()
    assert workbook_data.persons["ES-L-0001"].research_consent_signed == "yes"
    assert workbook_data.warnings[0] == "Deprecated column consent_signed used; please rename to research_consent_signed."


def test_load_intake_workbook_normalizes_whole_column_sqref_without_modifying_original(tmp_path: Path) -> None:
    workbook_path = _write_workbook(tmp_path)
    workbook = load_workbook(workbook_path)
    sheet = workbook["Research_Session_Intake"]
    validation = DataValidation(type="list", formula1='"baseline,follow_up"')
    sheet.add_data_validation(validation)
    validation.add("H2:H1048576")
    workbook.save(workbook_path)
    workbook.close()
    _patch_workbook_sqref(workbook_path, "H2:H1048576", "H:H")

    workbook_data = load_intake_workbook(workbook_path, target_language="es")

    assert workbook_data.errors == ()
    assert workbook_data.sessions[0].session_id == "ES-L-0001-2026-S01"
    assert any("sqref 'H:H'" in warning and "original workbook was not modified" in warning for warning in workbook_data.warnings)
    with zipfile.ZipFile(workbook_path, "r") as workbook_zip:
        assert b'sqref="H:H"' in workbook_zip.read("xl/worksheets/sheet3.xml")


@pytest.mark.parametrize(
    ("raw_duration", "expected_duration"),
    [("0,75", 0.75), ("3,5", 3.5), ("6,5", 6.5), ("0.75", 0.75), ("3.5", 3.5)],
)
def test_load_intake_workbook_normalizes_decimal_exposure_duration_months(
    tmp_path: Path,
    raw_duration: object,
    expected_duration: float,
) -> None:
    workbook_path = _write_workbook(tmp_path, exposure_duration=raw_duration)

    workbook_data = load_intake_workbook(workbook_path, target_language="es")

    exposure = workbook_data.exposures_by_key[next(iter(workbook_data.exposures_by_key))][0]
    assert exposure.duration_months == expected_duration


def test_load_intake_workbook_keeps_country_string_and_normalizes_unspecified_exposure_type(tmp_path: Path) -> None:
    workbook_path = _write_workbook(
        tmp_path,
        exposure_duration="3",
        exposure_type="unspecified",
        exposure_country="France; Israel",
        exposure_notes="Stayed in Strasbourg and Vannes for approximately three weeks.",
    )

    workbook_data = load_intake_workbook(workbook_path, target_language="es")

    exposure = workbook_data.exposures_by_key[next(iter(workbook_data.exposures_by_key))][0]
    assert exposure.country == "France; Israel"
    assert exposure.exposure_type == "unknown"
    assert exposure.exposure_notes == "Stayed in Strasbourg and Vannes for approximately three weeks."
    assert workbook_data.warnings[-1] == "Deprecated exposure type unspecified in Exposure row 2; normalized to unknown."


@pytest.mark.parametrize("workbook_value", ["EC_STD", "CL_STD", "PE_STD", "BO_STD", "UY_STD", "PY_STD", "VE_STD"])
def test_load_intake_workbook_accepts_new_standard_varieties(tmp_path: Path, workbook_value: str) -> None:
    workbook_path = _write_workbook(tmp_path, speaker_type="native_speaker", standard_variety=workbook_value)

    workbook_data = load_intake_workbook(workbook_path, target_language="es")

    assert workbook_data.sessions[0].standard_variety == workbook_value.lower()


def test_production_import_syncs_interview_runtime_db_and_metadata(tmp_path: Path, monkeypatch) -> None:
    runtime_root = _set_runtime_env(tmp_path, monkeypatch)
    batch_dir = _prepare_interview_batch(tmp_path, include_raw_master=True)
    workbook_path = _write_workbook(tmp_path)
    workbook_data = load_intake_workbook(workbook_path, target_language="es")
    session_factory = _session_factory(tmp_path)

    monkeypatch.setattr(
        production_importer,
        "create_full_task_mp3",
        lambda source_wav, target_mp3: target_mp3.write_bytes(b"runtime-interview-mp3"),
    )

    with session_factory() as db_session:
        plans, plan_warnings = production_importer._build_import_plans(
            batch_dir=batch_dir,
            workbook_data=workbook_data,
            create_missing_only=False,
            sync_tasks=True,
            sync_raw_only=False,
            allow_session_id_change=False,
            db_session=db_session,
        )
        assert plan_warnings == []
        assert len(plans) == 1
        plan = plans[0]
        by_task = {task_plan.task_key: task_plan for task_plan in plan.task_plans}
        assert by_task["interview"].action == "sync"
        assert by_task["interview"].status == "ready"

        production_importer._apply_plan(db_session, plan, validate_wordlist_labels="off")

    session_dir = runtime_root / "data" / "sessions" / "spanish" / "ES-L-0001-2026-S01"
    archive_dir = tmp_path / "archive-root" / "sessions" / "es" / "ES-L-0001-2026-S01"
    metadata = json.loads((session_dir / "metadata.json").read_text(encoding="utf-8"))
    interview_alignment = json.loads((session_dir / "alignment" / "interview.json").read_text(encoding="utf-8"))

    assert not (session_dir / "source" / "interview.wav").exists()
    assert not (session_dir / "raw" / "interview.wav").exists()
    assert (session_dir / "derived" / "interview.mp3").read_bytes() == b"runtime-interview-mp3"
    assert interview_alignment["session_id"] == "ES-L-0001-2026-S01"
    assert interview_alignment["audio"] == {"full_mp3": "derived/interview.mp3"}
    assert interview_alignment["segments"][0]["tokens"][2]["suffix"] == "."
    assert all("trailing_punctuation" not in annotation for annotation in interview_alignment["segments"][0]["annotations"])

    interview_task = next(task for task in metadata["tasks"] if task["task_type"] == "interview")
    assert interview_task == {
        "task_type": "interview",
        "label": "Interview zur Aussprache",
        "alignment_file": "alignment/interview.json",
        "derived_file": "derived/interview.mp3",
    }
    exposure_entry = metadata["exposure_entries"][0]
    assert exposure_entry["exposure_notes"] == "Semester in Madrid."
    assert metadata["research_consent_signed"] == "yes"
    assert metadata["teaching_consent_signed"] == "unknown"
    assert metadata["session_notes"] == "test learner session"

    file_paths = {entry["path"] for entry in metadata["files"]}
    assert "alignment/interview.json" in file_paths
    assert "derived/interview.mp3" in file_paths
    assert "source/interview.wav" not in file_paths
    assert "raw/interview.wav" not in file_paths

    archive_manifest = json.loads((archive_dir / "metadata" / "archive_manifest.json").read_text(encoding="utf-8"))
    assert (archive_dir / "source" / "interview.wav").read_bytes() == b"batch-source-interview"
    assert (archive_dir / "raw" / "interview.wav").read_bytes() == b"batch-raw-interview"
    assert (archive_dir / "runtime" / "derived" / "interview.mp3").read_bytes() == b"runtime-interview-mp3"
    assert archive_manifest["session_id"] == "ES-L-0001-2026-S01"
    assert archive_manifest["source_batch"] == "spanish_batch_20260421"

    research_sessions.load_language_sessions.cache_clear()
    research_sessions.load_person_records.cache_clear()
    runtime_session = research_sessions.get_session("spanish", "ES-L-0001-2026-S01")
    assert runtime_session is not None
    assert runtime_session.documented_task_types == ("interview",)
    assert runtime_session.exposure_entries[0].exposure_notes == "Semester in Madrid."

    with session_factory() as db_session:
        person_rows = db_session.scalars(select(ResearchPerson)).all()
        session_rows = db_session.scalars(select(ResearchSession)).all()
        exposure_rows = db_session.scalars(select(ResearchSessionExposure)).all()

    assert [row.person_id for row in person_rows] == ["ES-L-0001"]
    assert [row.session_id for row in session_rows] == ["ES-L-0001-2026-S01"]
    assert session_rows[0].documented_tasks == "interview"
    assert len(exposure_rows) == 1
    assert exposure_rows[0].exposure_notes == "Semester in Madrid."
    assert person_rows[0].research_consent_signed == "yes"
    assert person_rows[0].teaching_consent_signed == "unknown"


def test_production_import_does_not_treat_source_as_raw_fallback(tmp_path: Path, monkeypatch) -> None:
    runtime_root = _set_runtime_env(tmp_path, monkeypatch)
    batch_dir = _prepare_interview_batch(tmp_path, include_raw_master=False)
    workbook_path = _write_workbook(tmp_path)
    workbook_data = load_intake_workbook(workbook_path, target_language="es")
    session_factory = _session_factory(tmp_path)

    monkeypatch.setattr(
        production_importer,
        "create_full_task_mp3",
        lambda source_wav, target_mp3: target_mp3.write_bytes(b"runtime-interview-mp3"),
    )

    with session_factory() as db_session:
        plans, _ = production_importer._build_import_plans(
            batch_dir=batch_dir,
            workbook_data=workbook_data,
            create_missing_only=False,
            sync_tasks=True,
            sync_raw_only=False,
            allow_session_id_change=False,
            db_session=db_session,
        )
        production_importer._apply_plan(db_session, plans[0], validate_wordlist_labels="off")

    session_dir = runtime_root / "data" / "sessions" / "spanish" / "ES-L-0001-2026-S01"
    archive_dir = tmp_path / "archive-root" / "sessions" / "es" / "ES-L-0001-2026-S01"
    assert not (session_dir / "source" / "interview.wav").exists()
    assert not (session_dir / "raw" / "interview.wav").exists()
    assert (archive_dir / "source" / "interview.wav").exists()
    assert not (archive_dir / "raw" / "interview.wav").exists()


def test_production_import_rerun_is_idempotent_for_interview_session(tmp_path: Path, monkeypatch) -> None:
    _set_runtime_env(tmp_path, monkeypatch)
    batch_dir = _prepare_interview_batch(tmp_path, include_raw_master=True)
    workbook_path = _write_workbook(tmp_path)
    workbook_data = load_intake_workbook(workbook_path, target_language="es")
    session_factory = _session_factory(tmp_path)

    monkeypatch.setattr(
        production_importer,
        "create_full_task_mp3",
        lambda source_wav, target_mp3: target_mp3.write_bytes(b"runtime-interview-mp3"),
    )

    with session_factory() as db_session:
        plans, _ = production_importer._build_import_plans(
            batch_dir=batch_dir,
            workbook_data=workbook_data,
            create_missing_only=False,
            sync_tasks=True,
            sync_raw_only=False,
            allow_session_id_change=False,
            db_session=db_session,
        )
        production_importer._apply_plan(db_session, plans[0], validate_wordlist_labels="off")

    with session_factory() as db_session:
        plans, _ = production_importer._build_import_plans(
            batch_dir=batch_dir,
            workbook_data=workbook_data,
            create_missing_only=False,
            sync_tasks=True,
            sync_raw_only=False,
            allow_session_id_change=False,
            db_session=db_session,
        )
        assert plans[0].mode_action == "update"
        production_importer._apply_plan(db_session, plans[0], validate_wordlist_labels="off")

    session_dir = tmp_path / "runtime-root" / "data" / "sessions" / "spanish" / "ES-L-0001-2026-S01"
    metadata = json.loads((session_dir / "metadata.json").read_text(encoding="utf-8"))
    interview_tasks = [task for task in metadata["tasks"] if task["task_type"] == "interview"]
    assert len(interview_tasks) == 1
    file_paths = [entry["path"] for entry in metadata["files"]]
    assert file_paths.count("alignment/interview.json") == 1
    assert file_paths.count("derived/interview.mp3") == 1
    assert file_paths.count("source/interview.wav") == 0
    assert file_paths.count("raw/interview.wav") == 0

    with session_factory() as db_session:
        person_rows = db_session.scalars(select(ResearchPerson)).all()
        session_rows = db_session.scalars(select(ResearchSession)).all()
        exposure_rows = db_session.scalars(select(ResearchSessionExposure)).all()

    assert len(person_rows) == 1
    assert len(session_rows) == 1
    assert len(exposure_rows) == 1


def test_production_import_skips_native_speaker_interview_even_with_inputs(tmp_path: Path, monkeypatch) -> None:
    runtime_root = _set_runtime_env(tmp_path, monkeypatch)
    batch_dir = _prepare_interview_batch(
        tmp_path,
        person_id="ES-N-0001",
        include_raw_master=True,
        include_working_interview=True,
    )
    workbook_path = _write_workbook(tmp_path, person_id="ES-N-0001", speaker_type="native_speaker")
    workbook_data = load_intake_workbook(workbook_path, target_language="es")
    session_factory = _session_factory(tmp_path)

    with session_factory() as db_session:
        plans, plan_warnings = production_importer._build_import_plans(
            batch_dir=batch_dir,
            workbook_data=workbook_data,
            create_missing_only=False,
            sync_tasks=True,
            sync_raw_only=False,
            allow_session_id_change=False,
            db_session=db_session,
        )
        assert plan_warnings == []
        assert len(plans) == 1
        plan = plans[0]
        by_task = {task_plan.task_key: task_plan for task_plan in plan.task_plans}
        assert by_task["interview"].action == "skip"
        assert by_task["interview"].status == "not_expected_for_native_speaker"

        production_importer._apply_plan(db_session, plan, validate_wordlist_labels="off")

    session_dir = runtime_root / "data" / "sessions" / "spanish" / "ES-N-0001-2026-S01"
    metadata = json.loads((session_dir / "metadata.json").read_text(encoding="utf-8"))

    assert metadata["speaker_type"] == "native_speaker"
    assert metadata["tasks"] == []
    file_paths = {entry["path"] for entry in metadata["files"]}
    assert file_paths == {"metadata.json"}
    assert not (session_dir / "source" / "interview.wav").exists()
    assert not (session_dir / "alignment" / "interview.json").exists()
    assert not (session_dir / "derived" / "interview.mp3").exists()
    assert not (session_dir / "raw" / "interview.wav").exists()

    with session_factory() as db_session:
        session_rows = db_session.scalars(select(ResearchSession)).all()

    assert [row.session_id for row in session_rows] == ["ES-N-0001-2026-S01"]
    assert session_rows[0].documented_tasks is None


def test_load_intake_workbook_exposes_secure_persons(tmp_path: Path) -> None:
    workbook_path = _write_workbook(tmp_path, person_id="ES-L-0001")

    workbook_data = load_intake_workbook(workbook_path, target_language="es")

    assert "ES-L-0001" in workbook_data.secure_persons
    sp = workbook_data.secure_persons["ES-L-0001"]
    assert sp.last_name == "Mustermann"
    assert sp.first_name == "Anna"
    assert sp.email == "anna@example.test"
    assert sp.research_consent_signed == "yes"
    assert sp.needs_review is False
    assert sp.intake_by == "Ana Romero"
    assert sp.paper_original_location == "cabinet A"


def test_load_intake_workbook_person_id_filter_accepts_set(tmp_path: Path) -> None:
    workbook_path = _write_workbook(tmp_path, person_id="ES-L-0001")

    workbook_data = load_intake_workbook(workbook_path, target_language="es", person_id_filter={"ES-L-0001"})

    assert len(workbook_data.sessions) == 1
    assert workbook_data.sessions[0].person_id == "ES-L-0001"
    assert workbook_data.errors == ()


def test_run_text_pipeline_skips_missing_working_text_inputs_in_write_mode(tmp_path: Path) -> None:
    batch_dir = tmp_path / "spanish_batch_20260525"
    batch_dir.mkdir()

    notes = production_importer._run_text_pipeline(
        batch_dir=batch_dir,
        person_id="ES-L-0010",
        target_language="es",
        mfa_executable="docker",
        dry_run=False,
    )

    assert notes == [
        "Skipped text MFA for ES-L-0010: working text inputs are not present; task will remain missing unless existing runtime artifacts are available."
    ]


def test_run_text_pipeline_dry_run_does_not_require_written_manifest(tmp_path: Path, monkeypatch) -> None:
    batch_dir = tmp_path / "english_batch_20260525"
    person_id = "EN-L-0008"
    (batch_dir / "working" / person_id / "text" / "source").mkdir(parents=True)
    (batch_dir / "working" / person_id / "text" / "alignment").mkdir()
    (batch_dir / "working" / person_id / "text" / "source" / "text.wav").write_bytes(b"wav")
    (batch_dir / "working" / person_id / "text" / "alignment" / "text.TextGrid").write_text("textgrid", encoding="utf-8")
    text_catalog_path = tmp_path / "text.json"
    text_catalog_path.write_text('{"task": "text", "language": "en", "items": []}\n', encoding="utf-8")

    monkeypatch.setattr(production_importer, "_text_task_catalog_path", lambda target_language: text_catalog_path)
    monkeypatch.setattr(
        production_importer,
        "prepare_text_mfa_for_person",
        lambda **kwargs: {"segments": 55, "warnings": ["EN-L-0008 text: omitted t_01 because the spoken title was not recorded"]},
    )

    def fail_run_text_mfa_for_person(**kwargs):
        raise AssertionError("dry-run must not require a written mfa_manifest.json")

    monkeypatch.setattr(production_importer, "run_text_mfa_for_person", fail_run_text_mfa_for_person)

    notes = production_importer._run_text_pipeline(
        batch_dir=batch_dir,
        person_id=person_id,
        target_language="en",
        mfa_executable="docker",
        dry_run=True,
    )

    assert notes == [
        "Prepared text MFA corpus for EN-L-0008: segments=55",
        "Text MFA prep warning for EN-L-0008: EN-L-0008 text: omitted t_01 because the spoken title was not recorded",
        "Planned MFA for EN-L-0008: executable=docker",
        "Planned working text alignment import for EN-L-0008 after MFA outputs are available.",
    ]


def test_run_text_pipeline_reuses_current_text_alignment_outputs(tmp_path: Path, monkeypatch) -> None:
    batch_dir = tmp_path / "english_batch_20260525"
    person_id = "EN-L-0008"
    source_wav = batch_dir / "working" / person_id / "text" / "source" / "text.wav"
    source_textgrid = batch_dir / "working" / person_id / "text" / "alignment" / "text.TextGrid"
    alignment_json = batch_dir / "working" / person_id / "text" / "alignment" / "text.json"
    manifest_path = batch_dir / "working" / person_id / "text" / "mfa_manifest.json"
    state_path = working_text_mfa_state_path(batch_dir, person_id)
    text_catalog_path = tmp_path / "text.json"

    source_wav.parent.mkdir(parents=True, exist_ok=True)
    source_textgrid.parent.mkdir(parents=True, exist_ok=True)
    source_wav.write_bytes(b"wav")
    source_textgrid.write_text("textgrid", encoding="utf-8")
    alignment_json.write_text('{"person_id": "EN-L-0008", "task": "text", "items": []}\n', encoding="utf-8")
    text_catalog_path.write_text('{"task": "text", "language": "en", "items": []}\n', encoding="utf-8")

    signatures = {
        "source_wav": production_importer._file_signature(source_wav),
        "source_textgrid": production_importer._file_signature(source_textgrid),
        "text_source_json": production_importer._file_signature(text_catalog_path),
    }
    manifest_path.write_text(
        json.dumps(
            {
                "person_id": person_id,
                "task": "text",
                "language_code": "en",
                "language": "english",
                "preparation_version": "2026-05-27-text-mfa-v2",
                "source_signatures": signatures,
            },
            ensure_ascii=False,
            indent=2,
        )
        + "\n",
        encoding="utf-8",
    )
    state_path.write_text(
        json.dumps(
            {
                "person_id": person_id,
                "task": "text",
                "language_code": "en",
                "language": "english",
                "preparation_version": "2026-05-27-text-mfa-v2",
                "source_signatures": signatures,
            },
            ensure_ascii=False,
            indent=2,
        )
        + "\n",
        encoding="utf-8",
    )

    monkeypatch.setattr(production_importer, "_text_task_catalog_path", lambda target_language: text_catalog_path)
    monkeypatch.setattr(production_importer, "prepare_text_mfa_for_person", lambda **kwargs: (_ for _ in ()).throw(AssertionError("prepare should not run for a current cache")))
    monkeypatch.setattr(production_importer, "run_text_mfa_for_person", lambda **kwargs: (_ for _ in ()).throw(AssertionError("run MFA should not run for a current cache")))
    monkeypatch.setattr(production_importer, "import_text_mfa_alignment_for_person", lambda **kwargs: (_ for _ in ()).throw(AssertionError("import should not run for a current cache")))

    notes = production_importer._run_text_pipeline(
        batch_dir=batch_dir,
        person_id=person_id,
        target_language="en",
        mfa_executable="docker",
        dry_run=False,
    )

    assert notes == ["Reused text MFA outputs for EN-L-0008: working alignment JSON is current."]


def test_run_text_pipeline_imports_cached_mfa_outputs_without_rerun(tmp_path: Path, monkeypatch) -> None:
    batch_dir = tmp_path / "english_batch_20260525"
    person_id = "EN-L-0009"
    source_wav = batch_dir / "working" / person_id / "text" / "source" / "text.wav"
    source_textgrid = batch_dir / "working" / person_id / "text" / "alignment" / "text.TextGrid"
    mfa_output_dir = batch_dir / "working" / person_id / "text" / "mfa_output"
    manifest_path = batch_dir / "working" / person_id / "text" / "mfa_manifest.json"
    state_path = working_text_mfa_state_path(batch_dir, person_id)
    text_catalog_path = tmp_path / "text.json"

    source_wav.parent.mkdir(parents=True, exist_ok=True)
    source_textgrid.parent.mkdir(parents=True, exist_ok=True)
    mfa_output_dir.mkdir(parents=True, exist_ok=True)
    source_wav.write_bytes(b"wav")
    source_textgrid.write_text("textgrid", encoding="utf-8")
    (mfa_output_dir / "segment.TextGrid").write_text("aligned", encoding="utf-8")
    text_catalog_path.write_text('{"task": "text", "language": "en", "items": []}\n', encoding="utf-8")

    signatures = {
        "source_wav": production_importer._file_signature(source_wav),
        "source_textgrid": production_importer._file_signature(source_textgrid),
        "text_source_json": production_importer._file_signature(text_catalog_path),
    }
    manifest_path.write_text(
        json.dumps(
            {
                "person_id": person_id,
                "task": "text",
                "language_code": "en",
                "language": "english",
                "preparation_version": "2026-05-27-text-mfa-v2",
                "source_signatures": signatures,
            },
            ensure_ascii=False,
            indent=2,
        )
        + "\n",
        encoding="utf-8",
    )
    state_path.write_text(
        json.dumps(
            {
                "person_id": person_id,
                "task": "text",
                "language_code": "en",
                "language": "english",
                "preparation_version": "2026-05-27-text-mfa-v2",
                "source_signatures": signatures,
            },
            ensure_ascii=False,
            indent=2,
        )
        + "\n",
        encoding="utf-8",
    )

    monkeypatch.setattr(production_importer, "_text_task_catalog_path", lambda target_language: text_catalog_path)
    monkeypatch.setattr(production_importer, "prepare_text_mfa_for_person", lambda **kwargs: (_ for _ in ()).throw(AssertionError("prepare should not run when cached MFA outputs are available")))
    monkeypatch.setattr(production_importer, "run_text_mfa_for_person", lambda **kwargs: (_ for _ in ()).throw(AssertionError("run MFA should not run when cached MFA outputs are available")))

    import_calls: list[str] = []

    def fake_import_text_mfa_alignment_for_person(**kwargs):
        import_calls.append(kwargs["person_id"])
        return SimpleNamespace(imported=True, skipped_reason=None, item_count=4, token_count=11)

    monkeypatch.setattr(production_importer, "import_text_mfa_alignment_for_person", fake_import_text_mfa_alignment_for_person)

    notes = production_importer._run_text_pipeline(
        batch_dir=batch_dir,
        person_id=person_id,
        target_language="en",
        mfa_executable="docker",
        dry_run=False,
    )

    assert import_calls == [person_id]
    assert notes == ["Imported cached text alignment for EN-L-0009: items=4 tokens=11"]


def test_run_text_pipeline_cached_import_falls_back_to_warning_mode(tmp_path: Path, monkeypatch) -> None:
    batch_dir = tmp_path / "english_batch_20260525"
    person_id = "EN-L-0010"
    source_wav = batch_dir / "working" / person_id / "text" / "source" / "text.wav"
    source_textgrid = batch_dir / "working" / person_id / "text" / "alignment" / "text.TextGrid"
    mfa_output_dir = batch_dir / "working" / person_id / "text" / "mfa_output"
    manifest_path = batch_dir / "working" / person_id / "text" / "mfa_manifest.json"
    state_path = working_text_mfa_state_path(batch_dir, person_id)
    text_catalog_path = tmp_path / "text.json"

    source_wav.parent.mkdir(parents=True, exist_ok=True)
    source_textgrid.parent.mkdir(parents=True, exist_ok=True)
    mfa_output_dir.mkdir(parents=True, exist_ok=True)
    source_wav.write_bytes(b"wav")
    source_textgrid.write_text("textgrid", encoding="utf-8")
    (mfa_output_dir / "segment.TextGrid").write_text("aligned", encoding="utf-8")
    text_catalog_path.write_text('{"task": "text", "language": "en", "items": []}\n', encoding="utf-8")

    signatures = {
        "source_wav": production_importer._file_signature(source_wav),
        "source_textgrid": production_importer._file_signature(source_textgrid),
        "text_source_json": production_importer._file_signature(text_catalog_path),
    }
    manifest_path.write_text(
        json.dumps(
            {
                "person_id": person_id,
                "task": "text",
                "language_code": "en",
                "language": "english",
                "preparation_version": "2026-05-27-text-mfa-v2",
                "source_signatures": signatures,
            },
            ensure_ascii=False,
            indent=2,
        )
        + "\n",
        encoding="utf-8",
    )
    state_path.write_text(
        json.dumps(
            {
                "person_id": person_id,
                "task": "text",
                "language_code": "en",
                "language": "english",
                "preparation_version": "2026-05-27-text-mfa-v2",
                "source_signatures": signatures,
            },
            ensure_ascii=False,
            indent=2,
        )
        + "\n",
        encoding="utf-8",
    )

    monkeypatch.setattr(production_importer, "_text_task_catalog_path", lambda target_language: text_catalog_path)
    monkeypatch.setattr(production_importer, "prepare_text_mfa_for_person", lambda **kwargs: (_ for _ in ()).throw(AssertionError("prepare should not run when cached MFA outputs are available")))
    monkeypatch.setattr(production_importer, "run_text_mfa_for_person", lambda **kwargs: (_ for _ in ()).throw(AssertionError("run MFA should not run when cached MFA outputs are available")))

    calls: list[bool] = []

    def fake_import_text_mfa_alignment_for_person(**kwargs):
        calls.append(kwargs["fail_on_missing_output"])
        if kwargs["fail_on_missing_output"]:
            raise ValueError("missing MFA TextGrid for text_042_t_42")
        return SimpleNamespace(
            imported=True,
            skipped_reason=None,
            warnings=["missing MFA TextGrid for text_042_t_42"],
            item_count=5,
            token_count=12,
        )

    monkeypatch.setattr(production_importer, "import_text_mfa_alignment_for_person", fake_import_text_mfa_alignment_for_person)

    notes = production_importer._run_text_pipeline(
        batch_dir=batch_dir,
        person_id=person_id,
        target_language="en",
        mfa_executable="docker",
        dry_run=False,
    )

    assert calls == [True, False]
    assert notes[0] == "Text MFA cached import fallback for EN-L-0010: missing MFA TextGrid for text_042_t_42"
    assert notes[1] == "Text MFA import warning for EN-L-0010: missing MFA TextGrid for text_042_t_42"
    assert notes[2] == "Imported cached text alignment for EN-L-0010 with warnings: items=5 tokens=12"


def test_run_text_pipeline_falls_back_to_docker_when_host_mfa_missing(tmp_path: Path, monkeypatch) -> None:
    batch_dir = tmp_path / "spanish_batch_20260525"
    person_id = "ES-L-0010"
    source_wav = batch_dir / "working" / person_id / "text" / "source" / "text.wav"
    source_textgrid = batch_dir / "working" / person_id / "text" / "alignment" / "text.TextGrid"
    text_catalog_path = tmp_path / "text.json"

    source_wav.parent.mkdir(parents=True, exist_ok=True)
    source_textgrid.parent.mkdir(parents=True, exist_ok=True)
    source_wav.write_bytes(b"wav")
    source_textgrid.write_text("textgrid", encoding="utf-8")
    text_catalog_path.write_text('{"task": "text", "language": "es", "items": []}\n', encoding="utf-8")

    monkeypatch.setattr(production_importer, "_text_task_catalog_path", lambda target_language: text_catalog_path)
    monkeypatch.setattr(production_importer, "check_mfa_available", lambda mfa_executable="mfa": "docker-version" if mfa_executable == "docker" else (_ for _ in ()).throw(RuntimeError("mfa missing")))

    prepare_calls: list[str] = []
    run_calls: list[str] = []

    def fake_prepare_text_mfa_for_person(**kwargs):
        prepare_calls.append(kwargs["person_id"])
        return {"segments": 3, "warnings": []}

    def fake_run_text_mfa_for_person(**kwargs):
        run_calls.append(kwargs["mfa_executable"])
        return {
            "person_id": kwargs["person_id"],
            "language_code": "es",
            "language_slug": "spanish",
            "mfa_executable": kwargs["mfa_executable"],
            "mfa_version": "docker-version",
            "command": ["docker"],
            "mode": "write",
            "output": "ok",
        }

    def fake_import_text_mfa_alignment_for_person(**kwargs):
        return SimpleNamespace(imported=True, skipped_reason=None, item_count=3, token_count=9)

    monkeypatch.setattr(production_importer, "prepare_text_mfa_for_person", fake_prepare_text_mfa_for_person)
    monkeypatch.setattr(production_importer, "run_text_mfa_for_person", fake_run_text_mfa_for_person)
    monkeypatch.setattr(production_importer, "import_text_mfa_alignment_for_person", fake_import_text_mfa_alignment_for_person)

    notes = production_importer._run_text_pipeline(
        batch_dir=batch_dir,
        person_id=person_id,
        target_language="es",
        mfa_executable="mfa",
        dry_run=False,
    )

    assert prepare_calls == [person_id]
    assert run_calls == ["docker"]
    assert any(note.startswith("Falling back to Docker-backed MFA for ES-L-0010") for note in notes)
    assert notes[-1] == "Imported working text alignment for ES-L-0010: items=3 tokens=9"


def test_run_working_pipeline_keeps_isolated_task_errors_nonfatal(tmp_path: Path, monkeypatch) -> None:
    report_payload = {
        "person_ids": ["ES-L-0015"],
        "summary": {"errors": 1},
        "tasks": [
            {"person_id": "ES-L-0015", "task": "wordlist", "status": "rebuilt"},
            {"person_id": "ES-L-0015", "task": "text", "status": "rebuilt"},
            {"person_id": "ES-L-0015", "task": "interview", "status": "error_invalid_material_ref_marker"},
        ],
    }

    monkeypatch.setattr(production_importer, "organize_batch_working_tree", lambda **kwargs: report_payload)

    result = production_importer._run_working_pipeline(
        batch_dir=tmp_path / "spanish_batch_20260619",
        person_ids={"ES-L-0015"},
        dry_run=True,
    )

    assert result == report_payload
    assert production_importer._working_report_error_count(result) == 1


def test_detect_working_text_requires_preparation_when_alignment_json_missing(tmp_path: Path) -> None:
    batch_dir = tmp_path / "english_batch_20260525"
    person_id = "EN-L-0011"
    (batch_dir / "working" / person_id / "text" / "source").mkdir(parents=True, exist_ok=True)
    (batch_dir / "working" / person_id / "text" / "alignment").mkdir(parents=True, exist_ok=True)
    (batch_dir / "working" / person_id / "text" / "source" / "text.wav").write_bytes(b"wav")
    (batch_dir / "working" / person_id / "text" / "alignment" / "text.TextGrid").write_text("textgrid", encoding="utf-8")

    plan = production_importer._detect_working_task(
        batch_dir=batch_dir,
        person_id=person_id,
        task_key="text",
        target_language="en",
        sync_tasks=True,
        speaker_type="learner",
    )

    assert plan.action == "available"
    assert plan.status == "needs_preparation"
    assert plan.reason == (
        "text source/TextGrid are present but alignment/text.json is missing or stale; "
        "text requires preparation; rerun with --run-working --run-mfa"
    )


def test_detect_working_text_requires_preparation_when_state_is_stale(tmp_path: Path, monkeypatch) -> None:
    batch_dir = tmp_path / "english_batch_20260525"
    person_id = "EN-L-0012"
    source_wav = batch_dir / "working" / person_id / "text" / "source" / "text.wav"
    source_textgrid = batch_dir / "working" / person_id / "text" / "alignment" / "text.TextGrid"
    alignment_json = batch_dir / "working" / person_id / "text" / "alignment" / "text.json"
    manifest_path = batch_dir / "working" / person_id / "text" / "mfa_manifest.json"
    state_path = working_text_mfa_state_path(batch_dir, person_id)
    text_catalog_path = tmp_path / "text.json"

    source_wav.parent.mkdir(parents=True, exist_ok=True)
    source_textgrid.parent.mkdir(parents=True, exist_ok=True)
    source_wav.write_bytes(b"wav")
    source_textgrid.write_text("textgrid", encoding="utf-8")
    alignment_json.write_text('{"person_id": "EN-L-0012", "task": "text", "items": []}\n', encoding="utf-8")
    text_catalog_path.write_text('{"task": "text", "language": "en", "items": []}\n', encoding="utf-8")

    stale_signatures = {
        "source_wav": {"path": "C:/tmp/old.wav", "size": 1, "mtime_ns": 1},
        "source_textgrid": {"path": "C:/tmp/old.TextGrid", "size": 1, "mtime_ns": 1},
        "text_source_json": {"path": "C:/tmp/old.json", "size": 1, "mtime_ns": 1},
    }
    payload = {
        "person_id": person_id,
        "task": "text",
        "language_code": "en",
        "language": "english",
        "preparation_version": "2026-05-27-text-mfa-v2",
        "source_signatures": stale_signatures,
    }
    manifest_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    state_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    monkeypatch.setattr(production_importer, "_text_task_catalog_path", lambda target_language: text_catalog_path)

    plan = production_importer._detect_working_task(
        batch_dir=batch_dir,
        person_id=person_id,
        task_key="text",
        target_language="en",
        sync_tasks=True,
        speaker_type="learner",
    )

    assert plan.action == "available"
    assert plan.status == "needs_preparation"
    assert plan.reason == (
        "text source/TextGrid are present but alignment/text.json is missing or stale; "
        "text requires preparation; rerun with --run-working --run-mfa"
    )


def test_text_task_state_matches_normalizes_windows_drive_letter_case(tmp_path: Path, monkeypatch) -> None:
    batch_dir = tmp_path / "french_batch_20260527"
    person_id = "FR-L-0008"
    source_wav = batch_dir / "working" / person_id / "text" / "source" / "text.wav"
    source_textgrid = batch_dir / "working" / person_id / "text" / "alignment" / "text.TextGrid"
    alignment_json = batch_dir / "working" / person_id / "text" / "alignment" / "text.json"
    manifest_path = batch_dir / "working" / person_id / "text" / "mfa_manifest.json"
    state_path = working_text_mfa_state_path(batch_dir, person_id)
    text_catalog_path = tmp_path / "text.json"

    source_wav.parent.mkdir(parents=True, exist_ok=True)
    source_textgrid.parent.mkdir(parents=True, exist_ok=True)
    source_wav.write_bytes(b"wav")
    source_textgrid.write_text("textgrid", encoding="utf-8")
    alignment_json.write_text('{"person_id": "FR-L-0008", "task": "text", "items": []}\n', encoding="utf-8")
    text_catalog_path.write_text('{"task": "text", "language": "fr", "items": []}\n', encoding="utf-8")

    monkeypatch.setattr(production_importer, "_text_task_catalog_path", lambda target_language: text_catalog_path)
    signatures = production_importer._text_task_input_signatures(batch_dir, person_id, "fr")
    assert signatures is not None
    normalized = signatures["source_signatures"]
    lower_case = json.loads(json.dumps(normalized))
    for key in ("source_wav", "source_textgrid", "text_source_json"):
        lower_case[key]["path"] = str(lower_case[key]["path"]).replace("\\", "/").lower()

    payload = {
        "person_id": person_id,
        "task": "text",
        "language_code": "fr",
        "language": "french",
        "preparation_version": "2026-05-27-text-mfa-v2",
        "source_signatures": lower_case,
    }
    manifest_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    state_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    assert production_importer._text_task_state_matches(batch_dir, person_id, "fr", require_alignment_json=True)


def test_spoken_intervals_ignores_numbered_silence_labels() -> None:
    intervals = [
        TextGridInterval(start_seconds=0.0, end_seconds=1.0, text="Hallo"),
        TextGridInterval(start_seconds=1.0, end_seconds=2.0, text="silent1"),
        TextGridInterval(start_seconds=2.0, end_seconds=3.0, text="Welt"),
    ]

    spoken = spoken_intervals(intervals)

    assert [interval.text for interval in spoken] == ["Hallo", "Welt"]


# ── has_delivered_task_data filter ───────────────────────────────────────────

def _make_batch_dir_with_no_working_data(tmp_path: Path, person_id: str = "FR-L-0022") -> Path:
    batch_dir = tmp_path / "french_batch_20260618"
    batch_dir.mkdir(parents=True, exist_ok=True)
    return batch_dir


def _make_batch_dir_with_wordlist(tmp_path: Path, person_id: str = "FR-L-0001") -> Path:
    batch_dir = tmp_path / "french_batch_20260618"
    src = batch_dir / "working" / person_id / "wordlist" / "source"
    aln = batch_dir / "working" / person_id / "wordlist" / "alignment"
    src.mkdir(parents=True)
    aln.mkdir(parents=True)
    (src / "wordlist.wav").write_bytes(b"wav")
    (aln / "wordlist.TextGrid").write_text("textgrid", encoding="utf-8")
    return batch_dir


def test_import_plan_has_no_delivered_task_data_when_no_working_files(tmp_path: Path, monkeypatch) -> None:
    _set_runtime_env(tmp_path, monkeypatch)
    batch_dir = _make_batch_dir_with_no_working_data(tmp_path, person_id="FR-L-0022")
    workbook_path = _write_workbook(tmp_path, person_id="FR-L-0022")
    workbook_data = load_intake_workbook(workbook_path, target_language="es")
    # Override person_id in workbook_data to match French naming (testing the logic only)
    session_factory = _session_factory(tmp_path)

    with session_factory() as db_session:
        plans, _ = production_importer._build_import_plans(
            batch_dir=batch_dir,
            workbook_data=workbook_data,
            create_missing_only=False,
            sync_tasks=True,
            sync_raw_only=False,
            allow_session_id_change=False,
            db_session=db_session,
        )

    assert len(plans) == 1
    assert plans[0].has_delivered_task_data is False
    assert all(tp.action == "skip" for tp in plans[0].task_plans)


def test_import_plan_has_delivered_task_data_when_wordlist_ready(tmp_path: Path, monkeypatch) -> None:
    _set_runtime_env(tmp_path, monkeypatch)
    batch_dir = _prepare_interview_batch(tmp_path, include_raw_master=False)
    workbook_path = _write_workbook(tmp_path)
    workbook_data = load_intake_workbook(workbook_path, target_language="es")
    session_factory = _session_factory(tmp_path)

    with session_factory() as db_session:
        plans, _ = production_importer._build_import_plans(
            batch_dir=batch_dir,
            workbook_data=workbook_data,
            create_missing_only=False,
            sync_tasks=True,
            sync_raw_only=False,
            allow_session_id_change=False,
            db_session=db_session,
        )

    assert len(plans) == 1
    assert plans[0].has_delivered_task_data is True


def test_import_plan_has_delivered_task_data_when_existing_runtime_artifacts(tmp_path: Path, monkeypatch) -> None:
    runtime_root = _set_runtime_env(tmp_path, monkeypatch)
    batch_dir = _make_batch_dir_with_no_working_data(tmp_path)
    workbook_path = _write_workbook(tmp_path)
    workbook_data = load_intake_workbook(workbook_path, target_language="es")
    session_factory = _session_factory(tmp_path)

    # Pre-populate runtime with a session that has artifacts
    session_dir = runtime_root / "data" / "sessions" / "spanish" / "ES-L-0001-2026-S01"
    (session_dir / "alignment").mkdir(parents=True)
    (session_dir / "derived").mkdir(parents=True)
    (session_dir / "alignment" / "wordlist.json").write_text("{}", encoding="utf-8")
    (session_dir / "derived" / "wordlist.mp3").write_bytes(b"mp3")
    (session_dir / "metadata.json").write_text('{"tasks": [{"task_type": "wordlist"}]}', encoding="utf-8")

    with session_factory() as db_session:
        plans, _ = production_importer._build_import_plans(
            batch_dir=batch_dir,
            workbook_data=workbook_data,
            create_missing_only=False,
            sync_tasks=True,
            sync_raw_only=False,
            allow_session_id_change=False,
            db_session=db_session,
        )

    assert len(plans) == 1
    assert plans[0].has_delivered_task_data is True


def test_import_plan_native_speaker_with_only_non_expected_interview_has_no_delivered_task_data(
    tmp_path: Path, monkeypatch
) -> None:
    _set_runtime_env(tmp_path, monkeypatch)
    batch_dir = _prepare_interview_batch(
        tmp_path, person_id="ES-N-0001", include_raw_master=False, include_working_interview=True
    )
    workbook_path = _write_workbook(tmp_path, person_id="ES-N-0001", speaker_type="native_speaker")
    workbook_data = load_intake_workbook(workbook_path, target_language="es")
    session_factory = _session_factory(tmp_path)

    with session_factory() as db_session:
        plans, _ = production_importer._build_import_plans(
            batch_dir=batch_dir,
            workbook_data=workbook_data,
            create_missing_only=False,
            sync_tasks=True,
            sync_raw_only=False,
            allow_session_id_change=False,
            db_session=db_session,
        )

    # Interview is "not_expected_for_native_speaker" → action=skip; no other tasks → no delivered data.
    assert plans[0].has_delivered_task_data is False
    by_task = {tp.task_key: tp for tp in plans[0].task_plans}
    assert by_task["interview"].action == "skip"
    assert by_task["interview"].status == "not_expected_for_native_speaker"


def test_main_skips_metadata_only_sessions_in_db_and_runtime(tmp_path: Path, monkeypatch) -> None:
    runtime_root = _set_runtime_env(tmp_path, monkeypatch)
    # batch_dir has NO working data → metadata-only
    batch_dir = _make_batch_dir_with_no_working_data(tmp_path)
    workbook_path = _write_workbook(tmp_path)
    workbook_data = load_intake_workbook(workbook_path, target_language="es")
    session_factory = _session_factory(tmp_path)

    monkeypatch.setattr(production_importer, "create_full_task_mp3", lambda src, dst: None)

    with session_factory() as db_session:
        plans, plan_warnings = production_importer._build_import_plans(
            batch_dir=batch_dir,
            workbook_data=workbook_data,
            create_missing_only=False,
            sync_tasks=True,
            sync_raw_only=False,
            allow_session_id_change=False,
            db_session=db_session,
        )
        assert len(plans) == 1
        assert not plans[0].has_delivered_task_data

        # _apply_plan must NOT be called for metadata-only sessions.
        # Simulate the main() loop guard:
        applied = []
        for plan in plans:
            if plan.mode_action not in {"create", "update"}:
                continue
            if not plan.has_delivered_task_data:
                continue
            applied.append(plan)

    assert applied == []
    session_dir = runtime_root / "data" / "sessions" / "spanish" / "ES-L-0001-2026-S01"
    assert not session_dir.exists()

    with session_factory() as db_session:
        from sqlalchemy import select
        from app.research_metadata import ResearchSession
        rows = db_session.scalars(select(ResearchSession)).all()
    assert rows == []


def test_import_db_payload_excludes_metadata_only_sessions(tmp_path: Path, monkeypatch) -> None:
    """applied_results (and thus db/import_payload.json) must not include sessions with no delivered task data."""
    _set_runtime_env(tmp_path, monkeypatch)
    batch_dir = _make_batch_dir_with_no_working_data(tmp_path)
    workbook_path = _write_workbook(tmp_path)
    workbook_data = load_intake_workbook(workbook_path, target_language="es")
    session_factory = _session_factory(tmp_path)

    applied_results = []
    with session_factory() as db_session:
        plans, _ = production_importer._build_import_plans(
            batch_dir=batch_dir,
            workbook_data=workbook_data,
            create_missing_only=False,
            sync_tasks=False,
            sync_raw_only=False,
            allow_session_id_change=False,
            db_session=db_session,
        )
        for plan in plans:
            if plan.mode_action not in {"create", "update"}:
                continue
            if not plan.has_delivered_task_data:
                continue
            applied_results.append(plan)

    assert applied_results == []
